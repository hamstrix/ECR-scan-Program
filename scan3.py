#PROGRAM DE SCANARE A CASELOR
numcase = input ("Introduceti numarul total de case din comanda :")
#Mai intai generam un fisier cu numele Distribuitorului
nume_doc= open("C:\\users\\tehnic 7\\Desktop\\program scanare\\provizoriu.txt", "w")
f = []
l = []
s1=set()
s2=set()
#Cream o bucla While true pentru a avea continuitate
while True:
    
    n = input("\nSelecteaza culoarea casei (a/n): \n")
    if n=="n":
     
        while True:
            m=input("\nIntroduceti seria casei: \n")
            
            if m.startswith("DB42") :
                p= "DP25 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB41") :
                p= "DP150 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB44") :
                p="WP50 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB43"):
                p="WP500 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB47") :
                p="FP700 NEGRU "+m+ "\n"
                l.append(p)

            elif m.startswith("DB45"):
                p="DP05 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB48") :
                p="FP800 NEGRU "+m+"\n"
                l.append(p)
               
            elif m.startswith("DB49") :
                p="FP650 NEGRU "+m+"\n"
                l.append(p)
                
            elif m.startswith("DB46") :
                p="DP25X NEGRU "+m+"\n"
                l.append(p)

            elif m.startswith("stop"):
                sorted(l)
                sorted(f)
                s1 = set(l)
                s2 = set(f)
                break
            else :
                
                continue
    if n == "a":
     
        while True:
            m=input("\nIntroduceti seria casei: \n")
            
            if m.startswith("DB42") :
                p="DP25 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB41") :
                p="DP150 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB44") :
                p="WP50 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB43"):
                p="WP500 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB47") :
                p="FP700 ALB "+m+"\n"
                f.append(p)
            elif m.startswith("DB45"):
                p="DP05 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB48") :
                p="FP800 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB49") :
                p="FP650 ALB "+m+"\n"
                f.append(p)
                
            elif m.startswith("DB46") :
                p="DP25X ALB "+m+"\n"
                f.append(p)

            elif m.startswith("stop"):
                sorted(l)
                sorted(f)
                s1 = set(l)
                s2 = set(f)
                break
            else :
                
                continue
    
    inapoi = input("\nSelectam alta culoare?(y/n) : \n")
    if inapoi=="y" :
        continue
       
     #introducem doua conditii if pentru a vedea daca numarul de serii scanate sunt egale cu numarul de case introduse

    elif int(len(s1.union(s2))) < int(numcase):

        print ("\n\nNumarul caselor scanate este mai mic ca numarul necesar de case din comanda\n")
        print ("Verificati seriile caselor scanate :\n")
        for row in (s1.union(s2)) :
            print (row)

        continue
        
    elif int(len(s1.union(s2))) > int(numcase) :

        print ("\n\nNumarul caselor scanate este mai mare ca numarul necesar de case din comanda\n")
        print (" Verificati seriile caselor scanate :\n")
        for row in (s1.union(s2)) :
            print (row)
        nume_doc.writelines("Model Culoare Serie \n")
        nume_doc.writelines(s1.union(s2))
        break

    else:
        nume_doc.writelines("Model Culoare Serie \n")
        nume_doc.writelines(s1.union(s2))
        break


    
    

nume_doc.close()

from openpyxl import Workbook


#deschidem fisierul provizoriu
document  = open("C:\\users\\tehnic 7\\Desktop\\program scanare\\provizoriu.txt")
#deschidem o lista in care vor fi introduse toate intrarile din fisier
intrari=[]
#ne asiguram ca parcurgem documentul de la inceput
document.seek(0)
#impartim fiecare rand prin " " 

for intrare in document.readlines():
    intrari.append(intrare.rstrip("\n").split(" "))
#printam lista intrarilor
print(intrari)
#Deschidem un workbook nou

tabel = Workbook()
path = "C:\\Users\\Tehnic 7\\Desktop\\vanzari\\"+ input("Introduceti numele Distribuitorului:\n")+".xlsx"
tabel.save(path)
w1 = tabel.create_sheet("DP25 NEGRE",1)
w2 = tabel.create_sheet("DP150 NEGRE",3)
w3 = tabel.create_sheet("DP25 X",2)
w4 = tabel.create_sheet("WP50 NEGRE",5)
w5 = tabel.create_sheet("DP05 NEGRE",7)
w6 = tabel.create_sheet("WP500",9)
w7 = tabel.create_sheet("FP700",10)
w8 = tabel.create_sheet("FP800",11)
w9 = tabel.create_sheet("FP650",12)
w10 = tabel.create_sheet("DP25 ALBE",0)
w11 = tabel.create_sheet("DP150 ALBE",4)
w12 = tabel.create_sheet("WP50 ALBE",6)
w13 = tabel.create_sheet("DP05 ALBE",8)





for row in intrari:

    if "Model" in row or "DP25" in row and "NEGRU" in row:
        w1 = tabel["DP25 NEGRE"]
        w1.append(row)
        w1.auto_filter.ref = "A1:C2000"
        w1.auto_filter.add_filter_column(0, ["DP25"])
        w1.auto_filter.add_sort_condition("C2:C2000")
        w1.column_dimensions["A"].width = 18
        w1.column_dimensions["B"].width = 18
        w1.column_dimensions["C"].width = 20

    if "Model" in row or "DP150" in row and "NEGRU" in row:
        w2 = tabel["DP150 NEGRE"]
        w2.append(row)
        w2.auto_filter.ref = "A1:C2000"
        w2.auto_filter.add_filter_column(0, ["DP150"])
        w2.auto_filter.add_sort_condition("C2:C2000")
        w2.column_dimensions["A"].width = 18
        w2.column_dimensions["B"].width = 18
        w2.column_dimensions["C"].width = 20

    if "Model" in row or "DP25X" in row:
        w3 = tabel["DP25 X"]
        w3.append(row)
        w3.auto_filter.ref = "A1:C2000"
        w3.auto_filter.add_filter_column(0, ["DP25X"])
        w3.auto_filter.add_sort_condition("C2:C2000")
        w3.column_dimensions["A"].width = 18
        w3.column_dimensions["B"].width = 18
        w3.column_dimensions["C"].width = 20

    if "Model" in row or "WP50" in row and "NEGRU" in row:
        w4 = tabel["WP50 NEGRE"]
        w4.append(row)
        w4.auto_filter.ref = "A1:C2000"
        w4.auto_filter.add_filter_column(0, ["WP50"])
        w4.auto_filter.add_sort_condition("C2:C2000")
        w4.column_dimensions["A"].width = 18
        w4.column_dimensions["B"].width = 18
        w4.column_dimensions["C"].width = 20

    if "Model" in row or "DP05" in row and "NEGRU" in row:
        w5 = tabel["DP05 NEGRE"]
        w5.append(row)
        w5.auto_filter.ref = "A1:C2000"
        w5.auto_filter.add_filter_column(0, ["DP05"])
        w5.auto_filter.add_sort_condition("C2:C2000")
        w5.column_dimensions["A"].width = 18
        w5.column_dimensions["B"].width = 18
        w5.column_dimensions["C"].width = 20

    if "Model" in row or "WP500" in row:
        w6 = tabel["WP500"]
        w6.append(row)
        w6.auto_filter.ref = "A1:C2000"
        w6.auto_filter.add_filter_column(0, ["WP500"])
        w6.auto_filter.add_sort_condition("C2:C2000")
        w6.column_dimensions["A"].width = 18
        w6.column_dimensions["B"].width = 18
        w6.column_dimensions["C"].width = 20

    if "Model" in row or "FP700" in row:
        w7 = tabel["FP700"]
        w7.append(row)
        w7.auto_filter.ref = "A1:C2000"
        w7.auto_filter.add_filter_column(0, ["FP700"])
        w7.auto_filter.add_sort_condition("C2:C2000")
        w7.column_dimensions["A"].width = 18
        w7.column_dimensions["B"].width = 18
        w7.column_dimensions["C"].width = 20

    if "Model" in row or "FP800" in row:
        w8 = tabel["FP800"]
        w8.append(row)
        w8.auto_filter.ref = "A1:C2000"
        w8.auto_filter.add_filter_column(0, ["FP800"])
        w8.auto_filter.add_sort_condition("C2:C2000")
        w8.column_dimensions["A"].width = 18
        w8.column_dimensions["B"].width = 18
        w8.column_dimensions["C"].width = 20

    if "Model" in row or "FP650" in row:
        w9 = tabel["FP650"]
        w9.append(row)
        w9.auto_filter.ref = "A1:C2000"
        w9.auto_filter.add_filter_column(0, ["FP650"])
        w9.auto_filter.add_sort_condition("C2:C2000")
        w9.column_dimensions["A"].width = 18
        w9.column_dimensions["B"].width = 18
        w9.column_dimensions["C"].width = 20

    if "Model" in row or "DP25" in row and "ALB" in row:
        w10 = tabel["DP25 ALBE"]
        w10.append(row)
        w10.auto_filter.ref = "A1:C2000"
        w10.auto_filter.add_filter_column(0, ["DP25"])
        w10.auto_filter.add_sort_condition("C2:C2000")
        w10.column_dimensions["A"].width = 18
        w10.column_dimensions["B"].width = 18
        w10.column_dimensions["C"].width = 20

    if "Model" in row or "DP150" in row and "ALB" in row:
        w11 = tabel["DP150 ALBE"]
        w11.append(row)
        w11.auto_filter.ref = "A1:C2000"
        w11.auto_filter.add_filter_column(0, ["DP150"])
        w11.auto_filter.add_sort_condition("C2:C2000")
        w11.column_dimensions["A"].width = 18
        w11.column_dimensions["B"].width = 18
        w11.column_dimensions["C"].width = 20

    if "Model" in row or "WP50" in row and "ALB" in row:
        w12 = tabel["WP50 ALBE"]
        w12.append(row)
        w12.auto_filter.ref = "A1:C2000"
        w12.auto_filter.add_filter_column(0, ["WP50"])
        w12.auto_filter.add_sort_condition("C2:C2000")
        w12.column_dimensions["A"].width = 18
        w12.column_dimensions["B"].width = 18
        w12.column_dimensions["C"].width = 20

    if "Model" in row or "DP05" in row and "ALB" in row:
        w13 = tabel["DP05 ALBE"]
        w13.append(row)
        w13.auto_filter.ref = "A1:C2000"
        w13.auto_filter.add_filter_column(0, ["DP05"])
        w13.auto_filter.add_sort_condition("C2:C2000")
        w13.column_dimensions["A"].width = 18
        w13.column_dimensions["B"].width = 18
        w13.column_dimensions["C"].width = 20

del tabel["Sheet"]
tabel.save(path)
document.close()
tabel.close()